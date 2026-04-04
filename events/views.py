from django.views.decorators.http import require_GET
from datetime import timedelta, datetime, time
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth import logout, authenticate, login
from django.http import JsonResponse, HttpResponse
from django.http import Http404
from .models import Convention, ConventionDay, Panel, Tag, PanelHost, Room, PanelTag, PanelHostOrder
from .forms import ConventionForm, ConventionDayForm, PanelForm, PanelHostForm, TagForm, CSVImportForm, XLSXImportForm
import icalendar
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Prefetch
import csv
import math
import openpyxl
import base64

def is_admin(user):
    return user.is_staff


def schedule(request):
    convention = Convention.objects.first()
    if convention:
        return redirect('events:convention_detail', pk=convention.pk)
    return render(request, 'events/schedule.html', {
        'conventions': [],
        'current_convention_name': 'FurConnect',
    })


@require_GET
def privacy_policy(request):
    return render(request, 'events/privacy_policy.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been successfully logged out.')
    return redirect('events:schedule')

    # Get all hosts for the convention

def convention_detail(request, pk):
    convention = get_object_or_404(Convention, pk=pk)
    current_convention_name = convention.name if convention else 'FurConnect'

    # Get all days for the convention
    days = convention.days.all().order_by('date')

    # Get all unique tags and rooms for the convention
    unique_tags = Tag.objects.filter(panels__convention_day__convention=convention).distinct().order_by('name')
    unique_rooms = Room.objects.filter(convention=convention).order_by('name')

    # Get all hosts for the convention
    convention_hosts = PanelHost.objects.filter(panels__convention_day__convention=convention).distinct().order_by('name')
        
    # Initialize dictionary to store panels grouped by day and time
    panels_by_display_time = {}
    
    # Process each day
    for day in days:
        # Sort panels for the current day by start time
        sorted_panels = day.panels.all().order_by('start_time')

        # Initialize grouping for the current day
        panels_by_display_time[day.date] = {}

        # Group panels by exact start time
        for panel in sorted_panels:
            if panel.cancelled:
                continue
            start_time = panel.start_time

            # Get hosts ordered by their priority in PanelHostOrder
            panel.ordered_hosts = list(panel.host.all().order_by('panelhostorder__priority'))
            
            # Get tags ordered by their priority in PanelTag
            panel.ordered_tags = list(panel.tags.all().order_by('paneltag__priority'))

            # Add panel to the grouped dictionary
            if start_time not in panels_by_display_time[day.date]:
                panels_by_display_time[day.date][start_time] = []

            panels_by_display_time[day.date][start_time].append(panel)

    # Sort the days for displaying in the template
    sorted_display_days = sorted(panels_by_display_time.keys())
    
    # Reconstruct days structure with panels grouped by display start time
    display_days_with_panels = []
    for day_date in sorted_display_days:
        day_obj = days.get(date=day_date)
        display_day = {
            'original_day_obj': day_obj,
            'panels_by_time': []
        }
        
        # Sort times for this day
        sorted_times = sorted(panels_by_display_time[day_date].keys())
        
        # Add each time group to the day
        for slot_time in sorted_times:
            display_day['panels_by_time'].append({
                'start_time': slot_time,
                'panels': panels_by_display_time[day_date][slot_time]
            })
        
        display_days_with_panels.append(display_day)

    # Build a 2D grid matrix (hourly slots x rooms) for grid view, with rowspan for multi-hour panels
    days_matrix = []
    for display_day in display_days_with_panels:
        # Find rooms used for this day
        rooms_used = []
        panels_for_day = []
        for time_group in display_day['panels_by_time']:
            for panel in time_group['panels']:
                if panel.room and panel.room not in rooms_used:
                    rooms_used.append(panel.room)
                panels_for_day.append(panel)

        rooms_used = sorted(rooms_used, key=lambda r: r.name)

        # Determine daily span from first booked event to last event (30-min grid)
        day_date = display_day['original_day_obj'].date

        if panels_for_day:
            min_start_time = min(panel.start_time for panel in panels_for_day)
            max_end_time = max(panel.end_time for panel in panels_for_day)

            start_dt = datetime.combine(day_date, min_start_time)
            if start_dt.minute < 30:
                start_dt = start_dt.replace(minute=0, second=0, microsecond=0)
            else:
                start_dt = start_dt.replace(minute=30, second=0, microsecond=0)

            end_dt = datetime.combine(day_date, max_end_time)
            if end_dt <= start_dt:
                end_dt += timedelta(days=1)

            if end_dt.minute == 0 and end_dt.second == 0 and end_dt.microsecond == 0:
                pass
            elif end_dt.minute <= 30:
                end_dt = end_dt.replace(minute=30, second=0, microsecond=0)
            else:
                end_dt = (end_dt + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)

            # Extend to at least one hour so there is room to draw
            if end_dt <= start_dt + timedelta(hours=1):
                end_dt = start_dt + timedelta(hours=1)
        else:
            start_dt = datetime.combine(day_date, time(0, 0))
            end_dt = datetime.combine(day_date, time(23, 30)) + timedelta(minutes=30)

        total_half_hours = int((end_dt - start_dt).total_seconds() // (30 * 60))
        time_slots = [start_dt + timedelta(minutes=30 * i) for i in range(total_half_hours)]

        # Prepare panel placement map by room and slot-start
        panel_map = {}
        for panel in panels_for_day:
            panel_start = datetime.combine(day_date, panel.start_time)
            panel_end = datetime.combine(day_date, panel.end_time)
            if panel_end <= panel_start:
                panel_end += timedelta(days=1)

            # Round panel start down to 30-min slot
            start_minute = 0 if panel_start.minute < 30 else 30
            slot_start = panel_start.replace(minute=start_minute, second=0, microsecond=0)
            if slot_start < start_dt:
                slot_start = start_dt

            duration_minutes = (panel_end - panel_start).total_seconds() / 60.0
            rowspan = max(1, math.ceil(duration_minutes / 30.0))

            panel_map[(panel.room.id, slot_start)] = {
                'panel': panel,
                'rowspan': rowspan,
                'end_dt': slot_start + timedelta(minutes=30 * rowspan)
            }

        # Build matrix rows with cells containing panel, empty, or skipped (spanned)
        room_span_end = {room.id: start_dt for room in rooms_used}
        matrix_rows = []

        for slot in time_slots:
            row = {'time': slot.time(), 'cells': []}
            for room in rooms_used:
                if slot < room_span_end.get(room.id, start_dt):
                    row['cells'].append({'type': 'skip'})
                    continue

                panel_entry = panel_map.get((room.id, slot))
                if panel_entry:
                    row['cells'].append({
                        'type': 'panel',
                        'panel': panel_entry['panel'],
                        'rowspan': panel_entry['rowspan']
                    })
                    room_span_end[room.id] = slot + timedelta(minutes=30 * panel_entry['rowspan'])
                else:
                    row['cells'].append({'type': 'empty'})
            matrix_rows.append(row)

        days_matrix.append({'day': display_day['original_day_obj'], 'rows': matrix_rows, 'rooms': rooms_used})

    return render(request, 'events/convention_detail.html', {
        'convention': convention,
        'days': display_days_with_panels,
        'days_matrix': days_matrix,
        'unique_tags': unique_tags,
        'unique_rooms': unique_rooms,
        'convention_hosts': convention_hosts,
        'current_convention_name': current_convention_name,
        'is_staff': request.user.is_staff,
    })

@login_required
def convention_create(request):
        if request.method == 'POST':
            form = ConventionForm(request.POST)
            if form.is_valid():
                convention = form.save()
                messages.success(request, 'Convention created successfully!')
                return redirect('events:schedule')
        else:
            form = ConventionForm()
        
        # Fetch the current convention name for the title, or use a default
        current_convention = Convention.objects.first()
        current_convention_name = current_convention.name if current_convention else 'FurConnect'

        return render(request, 'events/convention_form.html', {
            'form': form,
            'action': 'Create',
            'current_convention_name': current_convention_name
            # 'states_by_country': STATES_BY_COUNTRY  # Commented out as states are handled by text input now
        })

@login_required
def convention_edit(request, pk):
    convention = get_object_or_404(Convention, pk=pk)
    # Fetch the current convention name
    current_convention = convention.name

    if request.method == 'POST':
        form = ConventionForm(request.POST, request.FILES, instance=convention)
        if form.is_valid():
            form.save()
            messages.success(request, 'Convention updated successfully!')
            # Use the pk from the URL arguments for the redirect
            return redirect('events:convention_detail', pk=pk)
    else:
        form = ConventionForm(instance=convention)
    # Use the current convention's name for the title
    current_convention_name = current_convention

    return render(request, 'events/convention_form.html', {
        'form': form,
        'action': 'Edit',
        'current_convention_name': current_convention_name
    })

@login_required
def panel_create(request, day_pk):
    # Get the ConventionDay using the provided day_pk from the URL
    convention_day_from_url = get_object_or_404(ConventionDay, pk=day_pk)
    # Get the convention associated with this day
    current_convention = convention_day_from_url.convention
    current_convention_name = current_convention.name if current_convention else 'FurConnect'

    # Instantiate the host form here so it's always available
    host_form = PanelHostForm()

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    print(f"Is AJAX request: {is_ajax}")

    if request.method == 'POST':
        # Pass the convention to the form to filter the convention_day queryset
        form = PanelForm(request.POST, convention=current_convention)
        if form.is_valid():
            panel = form.save(commit=False)
            # The convention_day is now selected via the form, no need to set it from URL pk
            panel.save()
            form.save_m2m() # Save ManyToMany data
            messages.success(request, 'Panel created successfully!')

            if is_ajax:
                 # Always return JSON for AJAX success
                 return JsonResponse({'success': True, 'redirect_url': redirect('events:convention_detail', pk=panel.convention_day.convention.pk).url})
            else:
                 # Redirect for non-AJAX success
                 return redirect('events:convention_detail', pk=panel.convention_day.convention.pk)
        else:
            # If form is invalid
            if is_ajax:
                 # Return JSON response with errors
                 return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            else:
                 # For non-AJAX, render the template with errors
                 return render(request, 'events/panel_form.html', {
                     'form': form,
                     'host_form': host_form,
                     'convention': current_convention,
                     'current_convention_name': current_convention_name,
                     'convention_pk': current_convention.pk
                 })
    else:
        # Pass the convention to the form to filter the convention_day queryset
        form = PanelForm(convention=current_convention)

    return render(request, 'events/panel_form.html', {
        'form': form,
        'host_form': host_form, # Pass the host form to the template
        'convention': current_convention, # Pass the convention object
        # 'date': convention_day.date, # No longer needed as day is selected in form
        'current_convention_name': current_convention_name,
        'convention_pk': current_convention.pk # Pass convention pk for redirect if needed
    })

@login_required
def panel_edit(request, pk):
    panel = get_object_or_404(Panel.objects.select_related('convention_day__convention').prefetch_related('tags', 'host'), pk=pk)
    # Store the convention_pk before saving, in case the object state changes
    convention_pk = panel.convention_day.convention.pk
    # Fetch the current convention name
    current_convention = panel.convention_day.convention
    current_convention_name = current_convention.name if current_convention else 'FurConnect'

    # Add ordered hosts and tags to the panel object
    panel.ordered_hosts = panel.get_ordered_hosts()
    panel.ordered_tags = panel.tags.all().order_by('paneltag__priority')

    if request.method == 'POST':
        form = PanelForm(request.POST, instance=panel)
        if form.is_valid():
            panel = form.save()
            messages.success(request, 'Panel updated successfully!')
            # Use the stored convention_pk for the redirect
            return redirect('events:convention_detail', pk=convention_pk)
    else:
        form = PanelForm(instance=panel)
        host_form = PanelHostForm() # Instantiate the host form
        tag_form = TagForm() # Instantiate the tag form
    return render(request, 'events/panel_form.html', {
        'form': form,
        'convention': panel.convention_day.convention,
        'date': panel.convention_day.date,
        'current_convention_name': current_convention_name,
        'convention_pk': panel.convention_day.convention.pk,
        'host_form': host_form,
        'tag_form': tag_form
    })

@login_required
def panel_delete(request, pk):
    panel = get_object_or_404(Panel, pk=pk)
    convention_pk = panel.convention_day.convention.pk
    
    if request.method == 'POST':
        panel.delete()
        messages.success(request, 'Panel deleted successfully!')
        return redirect('events:convention_detail', pk=convention_pk)
    
    return render(request, 'events/panel_confirm_delete.html', {
        'panel': panel,
        'current_convention_name': panel.convention_day.convention.name
    })

@login_required
@user_passes_test(is_admin)
def convention_delete(request, pk):
    convention = get_object_or_404(Convention, pk=pk)
    if request.method == 'POST':
        convention.delete()
        messages.success(request, 'Convention deleted successfully!')
        return redirect('events:schedule')
    # Optional: Add a GET request handler to show a confirmation page
    # else:
    #     return render(request, 'events/convention_confirm_delete.html', {'convention': convention})

def panel_detail_modal_view(request, pk):
    panel = get_object_or_404(Panel.objects.select_related('convention_day__convention').prefetch_related('tags', 'host'), pk=pk)
    # Add ordered hosts and tags to the panel object
    panel.ordered_hosts = list(panel.host.all().order_by('panelhostorder__priority'))
    panel.ordered_tags = list(panel.tags.all().order_by('paneltag__priority'))
    return render(request, 'events/panel_detail_modal.html', {'panel': panel})

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, 'Welcome back! You have been successfully logged in.')
            # Redirect to the schedule page on successful login
            return redirect('events:schedule')
        else:
            # Add an error message using Django's messages framework
            messages.error(request, 'Invalid username or password.')

    # Handle GET request or failed POST request by rendering the login template
    # If it was a failed POST, the messages framework will include the error.
    return render(request, 'events/login.html', {
        'current_convention_name': 'FurConnect',
        # You might need to pass a flag here if you want to show the register prompt
        # only when there are no users, but let's keep it simple for now.
        # 'show_register_prompt': True # You would determine this logic elsewhere
    })

@login_required
def add_panel_host_ajax(request):
    print("add_panel_host_ajax called")
    if request.method == 'POST':
        host_id = request.POST.get('host_id')
        image_base64 = request.POST.get('image_base64') # Get base64 string
        name = request.POST.get('name')

        print(f"Received POST data - host_id: {host_id}, name: {name}, image_base64 length: {len(image_base64) if image_base64 else 0}")

        if host_id:
            print(f"Attempting to update existing host with ID: {host_id}")
            # If host_id is provided, try to update the existing host
            try:
                host = PanelHost.objects.get(pk=host_id)
                print("Host found.")
                # Update host instance directly with base64 data
                host.name = request.POST.get('name', host.name) # Update name if provided
                print(f"Updating host name to: {host.name}")
                # Only update image if a new base64 string is provided
                if image_base64:
                    print("Updating host image with new base64 data.")
                    host.image = image_base64
                elif image_base64 == '' and host.image: # Handle explicit clearing of image
                    print("Clearing existing host image.")
                    host.image = None
                host.save()
                print("Host updated successfully.")
                return JsonResponse({
                    'success': True, 
                    'host': {
                        'id': host.pk, 
                        'name': host.name, 
                        'profile_picture': host.image if host.image else None
                    }
                })
            except PanelHost.DoesNotExist:
                print("Error: Host not found for update.")
                return JsonResponse({'success': False, 'error': 'Host not found.'}, status=404)
            except Exception as e:
                print(f"Error updating host: {e}")
                return JsonResponse({'success': False, 'error': str(e)}, status=400)
        else:
            print("Attempting to create new host.")
            # If no host_id, create a new host
            # Manually handle base64 image for creation
            host = PanelHost(name=name, image=image_base64 if image_base64 else None)
            try:
                 host.full_clean() # Validate model fields
                 host.save()
                 print("New host created successfully.")
                 return JsonResponse({
                     'success': True,
                     'host': {
                         'id': host.pk,
                         'name': host.name,
                         'profile_picture': host.image if host.image else None
                     }
                 })
            except ValidationError as e:
                print(f"Validation error creating host: {e.message_dict}")
                return JsonResponse({'success': False, 'errors': e.message_dict}, status=400)
            except Exception as e:
                print(f"Error creating host: {e}")
                return JsonResponse({'success': False, 'error': str(e)}, status=400)
    print("Invalid request method.")
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

@login_required
def add_tag_ajax(request):
    if request.method == 'POST':
        tag_id = request.POST.get('tag_id')
        if tag_id:
            # If tag_id is provided, try to update the existing tag
            try:
                tag = Tag.objects.get(pk=tag_id)
                form = TagForm(request.POST, instance=tag)
            except Tag.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Tag not found.'}, status=404)
        else:
            # If no tag_id, create a new tag
            form = TagForm(request.POST)

        if form.is_valid():
            tag = form.save()
            return JsonResponse({'success': True, 'tag': {'id': tag.pk, 'name': tag.name, 'color': tag.color}})
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

def panel_calendar(request, pk):
    panel = get_object_or_404(Panel.objects.select_related('convention_day__convention'), pk=pk)
    
    # Create calendar
    cal = icalendar.Calendar()
    cal.add('prodid', '-//FurConnect//Panel Calendar//EN')
    cal.add('version', '2.0')
    
    # Create event
    event = icalendar.Event()
    event.add('summary', panel.title)
    event.add('description', panel.description)
    event.add('location', f"{panel.convention_day.convention.name} - {panel.room}")
    
    # Set start and end times
    start_datetime = datetime.combine(panel.convention_day.date, panel.start_time)
    end_datetime = datetime.combine(panel.convention_day.date, panel.end_time)
    
    # Handle events that end after midnight
    if end_datetime < start_datetime:
        end_datetime += timedelta(days=1)
    
    event.add('dtstart', start_datetime)
    event.add('dtend', end_datetime)
    event.add('dtstamp', timezone.now())
    
    # Add event to calendar
    cal.add_component(event)
    
    # Create response
    response = HttpResponse(cal.to_ical(), content_type='text/calendar')
    response['Content-Disposition'] = f'attachment; filename="{panel.title}.ics"'
    
    return response

@login_required
def tag_edit(request, name):
    try:
        tag = Tag.objects.get(name__iexact=name)
        if request.method == 'POST':
            form = TagForm(request.POST, instance=tag)
            if form.is_valid():
                form.save()
                messages.success(request, 'Tag updated successfully!')
                # Redirect back to the convention detail page, or schedule if tag has no panels
                if tag.panels.exists():
                    return redirect('events:convention_detail', pk=tag.panels.first().convention_day.convention.pk)
                else:
                    # If the tag is not associated with any panels, redirect to the schedule
                    return redirect('events:schedule')
        else:
            form = TagForm(instance=tag)
        
        return render(request, 'events/tag_form.html', {
            'form': form,
            'tag': tag,
            'current_convention_name': 'FurConnect'
        })
    except Tag.DoesNotExist:
        messages.error(request, 'Tag not found.')
        return redirect('events:schedule')

    @login_required
    def host_edit(request, pk):
        host = get_object_or_404(PanelHost, pk=pk)
        if request.method == 'POST':
            form = PanelHostForm(request.POST, request.FILES, instance=host)
            if form.is_valid():
                # Handle image field
                if form.cleaned_data.get('image') is None:
                    host.image = None
                elif 'image' in request.FILES:
                    uploaded_file = request.FILES['image']
                    file_content = uploaded_file.read()
                    encoded = base64.b64encode(file_content).decode('utf-8')
                    content_type = uploaded_file.content_type
                    host.image = f"data:{content_type};base64,{encoded}"
                form.save()
                messages.success(request, 'Host updated successfully!')
                # Redirect back to the convention detail page, or schedule if host has no panels
                if host.panels.exists():
                    return redirect('events:convention_detail', pk=host.panels.first().convention_day.convention.pk)
                else:
                    # If the host is not associated with any panels, redirect to the schedule
                    return redirect('events:schedule')
        else:
            form = PanelHostForm(instance=host)
        
        return render(request, 'events/host_form.html', {
            'form': form,
            'host': host,
            'current_convention_name': 'FurConnect'
        })

@login_required
def delete_room_ajax(request, pk):
    """
    AJAX view to delete a single Room.
    """
    if request.method == 'POST':
        try:
            room = Room.objects.get(pk=pk)
            room.delete()
            return JsonResponse({'success': True})
        except Room.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Room not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

@login_required
def delete_host_ajax(request, pk):
    """
    AJAX view to delete a single PanelHost.
    """
    if request.method == 'POST':
        try:
            host = PanelHost.objects.get(pk=pk)
            host.delete()
            return JsonResponse({'success': True})
        except PanelHost.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Host not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

@login_required
def get_tag_details_ajax(request, pk):
    """
    AJAX view to get details of a single Tag.
    """
    try:
        tag = Tag.objects.get(pk=pk)
        return JsonResponse({
            'id': tag.pk,
            'name': tag.name,
            'color': tag.color
        })
    except Tag.DoesNotExist:
        return JsonResponse({'error': 'Tag not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def save_room_ajax(request):
    """
    AJAX view to save a new or existing Room.
    """
    if request.method == 'POST':
        try:
            room_id = request.POST.get('room_id')
            name = request.POST.get('name')
            convention_id = request.POST.get('convention_id')

            if not name or not convention_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Name and convention ID are required.'
                }, status=400)

            convention = get_object_or_404(Convention, pk=convention_id)

            if room_id:
                # Update existing room
                room = get_object_or_404(Room, pk=room_id)
                room.name = name
                room.save()
            else:
                # Create new room
                room = Room.objects.create(
                    name=name,
                    convention=convention
                )

            return JsonResponse({
                'success': True,
                'room': {
                    'id': room.pk,
                    'name': room.name
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method.'
    }, status=400)

@login_required
@user_passes_test(is_admin)
def manage_convention_items(request, pk):
    """Admin page to manage Rooms, Hosts, and Tags for a convention."""
    convention = get_object_or_404(Convention, pk=pk)
    
    rooms = Room.objects.filter(convention=convention).order_by('name')
    hosts = PanelHost.objects.filter(panels__convention_day__convention=convention).distinct().order_by('name')
    tags = Tag.objects.filter(panels__convention_day__convention=convention).distinct().order_by('name')
    
    return render(request, 'events/manage_convention_items.html', {
        'convention': convention,
        'rooms': rooms,
        'hosts': hosts,
        'tags': tags,
        'current_convention_name': convention.name,
    })


@login_required
@user_passes_test(is_admin)
def toggle_cancelled(request, pk):
    """Toggle the cancelled status of a panel."""
    
    try:
        panel = get_object_or_404(Panel, pk=pk)
        panel.cancelled = not panel.cancelled
        panel.save()
        
        # Check if the request wants JSON
        if request.headers.get('Accept') == 'application/json' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'cancelled': panel.cancelled,
                'message': 'Panel cancelled' if panel.cancelled else 'Panel uncancelled'
            })
        
        # For regular browser requests, redirect back to the convention detail page
        return redirect('events:convention_detail', pk=panel.convention_day.convention.pk)
        
    except Exception as e:
        if request.headers.get('Accept') == 'application/json' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': str(e)}, status=500)
        messages.error(request, f'Error toggling panel status: {str(e)}')
        return redirect('events:convention_detail', pk=panel.convention_day.convention.pk)

@login_required
@user_passes_test(is_admin)
def delete_tag_ajax(request, pk):
    """
    AJAX view to delete a single Tag.
    """
    if request.method == 'POST':
        try:
            tag = Tag.objects.get(pk=pk)
            tag.delete()
            return JsonResponse({'success': True})
        except Tag.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Tag not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

def get_host_details_ajax(request, pk):
    """
    AJAX view to get details of a single PanelHost.
    """
    try:
        host = PanelHost.objects.get(pk=pk)
        # Get all panels for this host
        panels = host.panels.all().select_related('convention_day', 'room').prefetch_related('tags')
        panels_data = []
        for panel in panels:
            tag_color = panel.tags.first().color if panel.tags.exists() else '#ffffff'
            panels_data.append({
                'id': panel.pk,
                'title': panel.title,
                'description': panel.description,
                'start_time': panel.start_time.strftime('%I:%M %p'),
                'end_time': panel.end_time.strftime('%I:%M %p'),
                'room_name': panel.room.name if panel.room else '',
                'tag_color': tag_color,
                'day_of_week': panel.convention_day.date.strftime('%A') if hasattr(panel, 'convention_day') and panel.convention_day and hasattr(panel.convention_day, 'date') and panel.convention_day.date else '',
                '_sort_date': panel.convention_day.date if hasattr(panel, 'convention_day') and panel.convention_day and hasattr(panel.convention_day, 'date') and panel.convention_day.date else None,
                '_sort_time': panel.start_time
            })
        # Sort by day (date), then by time
        panels_data.sort(key=lambda x: (x['_sort_date'], x['_sort_time']))
        for p in panels_data:
            p.pop('_sort_date', None)
            p.pop('_sort_time', None)
        
        return JsonResponse({
            'id': host.pk,
            'name': host.name,
            'profile_picture': host.image if host.image else host.get_initials_avatar(),
            'panels': panels_data,
            'panels_count': len(panels_data)
        })
    except PanelHost.DoesNotExist:
        return JsonResponse({'error': 'Host not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

def get_room_details_ajax(request, pk):
    """
    AJAX view to get details of a single Room.
    """
    try:
        room = Room.objects.get(pk=pk)
        return JsonResponse({
            'id': room.pk,
            'name': room.name
        })
    except Room.DoesNotExist:
        return JsonResponse({'error': 'Room not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def get_all_hosts_ajax(request):
    """
    AJAX view to get all PanelHosts for a given convention.
    Requires convention_id as a GET parameter.
    Optionally accepts panel_id to mark selected hosts.
    """
    convention_id = request.GET.get('convention_id')
    panel_id = request.GET.get('panel_id')

    if not convention_id:
        return JsonResponse({'error': 'convention_id is required.'}, status=400)

    try:
        # Filter hosts by convention and order by name
        hosts = PanelHost.objects.filter(panels__convention_day__convention__id=convention_id).distinct().order_by('name')

        # Determine selected hosts if panel_id is provided
        selected_host_ids = []
        if panel_id:
            try:
                panel = Panel.objects.get(pk=panel_id)
                selected_host_ids = list(panel.host.values_list('id', flat=True))
            except Panel.DoesNotExist:
                pass # Panel not found, no hosts are pre-selected

        hosts_data = []
        for host in hosts:
            hosts_data.append({
                'id': host.pk,
                'name': host.name,
                'profile_picture': host.image if host.image else host.get_initials_avatar(),
                'selected': host.id in selected_host_ids
            })

        return JsonResponse({'hosts': hosts_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def get_hosts_batch_ajax(request):
    """
    AJAX view to get host details for a batch of host IDs.
    Expects either:
    - ids=1,2,3
    - host_ids[]=1&host_ids[]=2
    """
    ids_param = request.GET.get('ids', '')
    host_ids = request.GET.getlist('host_ids[]')
    if not host_ids and ids_param:
        host_ids = [value.strip() for value in ids_param.split(',') if value.strip()]

    if not host_ids:
        return JsonResponse({'error': 'ids or host_ids[] is required.'}, status=400)

    try:
        hosts = PanelHost.objects.filter(pk__in=host_ids).prefetch_related('panels__convention_day', 'panels__room', 'panels__tags')
        hosts_data = []
        for host in hosts:
            panels_data = []
            for panel in host.panels.all():
                tag_color = panel.tags.first().color if panel.tags.exists() else '#ffffff'
                panels_data.append({
                    'id': panel.pk,
                    'title': panel.title,
                    'description': panel.description,
                    'date': panel.convention_day.date.strftime('%Y-%m-%d') if panel.convention_day else '',
                    'day_of_week': panel.convention_day.date.strftime('%A') if panel.convention_day and panel.convention_day.date else '',
                    'start_time': panel.start_time.strftime('%H:%M'),
                    'end_time': panel.end_time.strftime('%H:%M'),
                    'room': panel.room.name if panel.room else 'TBD',
                    'room_name': panel.room.name if panel.room else '',
                    'tag_color': tag_color,
                })

            hosts_data.append({
                'id': host.pk,
                'name': host.name,
                'profile_picture': host.image if host.image else host.get_initials_avatar(),
                'panels': panels_data,
                'panels_count': len(panels_data),
            })

        return JsonResponse({'hosts': hosts_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def host_edit(request, pk):
    host = get_object_or_404(PanelHost, pk=pk)
    if request.method == 'POST':
        form = PanelHostForm(request.POST, request.FILES, instance=host)
        if form.is_valid():
            if form.cleaned_data.get('image') is None:
                host.image = None
            elif 'image' in request.FILES:
                uploaded_file = request.FILES['image']
                file_content = uploaded_file.read()
                encoded = base64.b64encode(file_content).decode('utf-8')
                content_type = uploaded_file.content_type
                host.image = f"data:{content_type};base64,{encoded}"
            form.save()
            messages.success(request, 'Host updated successfully!')
            if host.panels.exists():
                return redirect('events:convention_detail', pk=host.panels.first().convention_day.convention.pk)
            return redirect('events:schedule')
    else:
        form = PanelHostForm(instance=host)

    return render(request, 'events/host_form.html', {
        'form': form,
        'host': host,
        'current_convention_name': 'FurConnect',
    })

@login_required
def get_all_rooms_ajax(request):
    """
    AJAX view to get all Rooms for a given convention.
    Requires convention_id as a GET parameter.
    """
    convention_id = request.GET.get('convention_id')

    if not convention_id:
        return JsonResponse({'error': 'convention_id is required.'}, status=400)

    try:
        rooms = Room.objects.filter(convention__id=convention_id)
        rooms_data = []
        for room in rooms:
            rooms_data.append({
                'id': room.pk,
                'name': room.name
            })
        return JsonResponse({'rooms': rooms_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def get_all_tags_ajax(request):
    """
    AJAX view to get all Tags for a given convention.
    Requires convention_id as a GET parameter.
    """
    convention_id = request.GET.get('convention_id')

    if not convention_id:
        return JsonResponse({'error': 'convention_id is required.'}, status=400)

    try:
        # Filter tags by convention (assuming Tag has a link to Convention, e.g., through Panels)
        # This might need adjustment based on your actual model relationships
        # Assuming a tag is related to a convention via a panel
        tags = Tag.objects.filter(panels__convention_day__convention__id=convention_id).distinct()

        tags_data = []
        for tag in tags:
            tags_data.append({
                'id': tag.pk,
                'name': tag.name,
                'color': tag.color
            })
        return JsonResponse({'tags': tags_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def reorder_tags_ajax(request, panel_id):
    """
    AJAX view to reorder tags for a panel.
    Expects a POST request with a list of tag IDs in the desired order.
    """
    if request.method == 'POST':
        try:
            panel = Panel.objects.get(pk=panel_id)
            tag_ids = request.POST.getlist('tag_ids[]')
            
            # Update priorities for each tag
            for index, tag_id in enumerate(tag_ids):
                PanelTag.objects.filter(panel=panel, tag_id=tag_id).update(priority=index)
            
            return JsonResponse({'success': True})
        except Panel.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Panel not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

@login_required
def reorder_hosts_ajax(request, panel_id):
    """
    AJAX view to reorder hosts for a panel.
    Expects a POST request with a list of host IDs in the desired order.
    """
    if request.method == 'POST':
        try:
            panel = Panel.objects.get(pk=panel_id)
            host_ids = request.POST.getlist('host_ids[]')
            
            # Update priorities for each host
            for index, host_id in enumerate(host_ids):
                PanelHostOrder.objects.filter(panel=panel, host_id=host_id).update(priority=index)
            
            return JsonResponse({'success': True})
        except Panel.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Panel not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

@login_required
def import_panels_csv(request, convention_pk):
    convention = get_object_or_404(Convention, pk=convention_pk)
    if request.method == 'POST':
        form = CSVImportForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = form.cleaned_data['csv_file']
            convention = form.cleaned_data['convention']

            # Read the CSV file and parse
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)

            # Map of possible column names to their canonical form
            column_mapping = {
                'title': ['title', 'Title'],
                'description': ['description', 'Description'],
                'date': ['date', 'Date'],
                'start_time': ['start time', 'Start Time', 'start_time'],
                'end_time': ['end time', 'End Time', 'end_time'],
                'room': ['room', 'Room'],
                'tags': ['tags', 'Tags'],
                'hosts': ['hosts', 'Hosts']
            }

            # Create a mapping of actual column names to canonical names
            actual_to_canonical = {}
            for canonical, possible_names in column_mapping.items():
                for name in possible_names:
                    if name in reader.fieldnames:
                        actual_to_canonical[name] = canonical
                        break

            # Check for missing required columns
            required_columns = ['title', 'description', 'date', 'start_time', 'end_time', 'room']
            missing_columns = [col for col in required_columns if col not in actual_to_canonical.values()]

            if missing_columns:
                messages.error(request, f"Missing required columns in CSV file: {', '.join(missing_columns)}")
                messages.info(request, "Required columns are: Title, Description, Date, Start Time, End Time, Room")
                messages.info(request, "Optional columns are: Tags, Hosts")
                return render(request, 'events/import_panels.html', {
                    'form': form,
                    'convention': convention,
                    'current_convention_name': convention.name
                })

            success_count = 0
            error_count = 0
            errors = []

            # Define possible date formats
            date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']

            for row in reader:
                with transaction.atomic():
                    try:
                        # Map the row data to canonical column names
                        mapped_row = {}
                        for actual_name, value in row.items():
                            if actual_name in actual_to_canonical:
                                mapped_row[actual_to_canonical[actual_name]] = value

                        # Clean and validate date value
                        date_str = mapped_row.get('date', '').strip()
                        if not date_str:
                            raise ValueError("Empty date value")

                        parsed_date = None
                        for date_format in date_formats:
                            try:
                                parsed_date = datetime.strptime(date_str, date_format).date()
                                break
                            except ValueError:
                                continue

                        if parsed_date is None:
                            raise ValueError(
                                f"Invalid date format: '{date_str}'. Supported formats are: YYYY-MM-DD, MM/DD/YYYY, DD/MM/YYYY, YYYY/MM/DD"
                            )

                        # Parse times
                        try:
                            start_time = datetime.strptime(mapped_row.get('start_time', '').strip(), '%H:%M').time()
                        except ValueError:
                            raise ValueError(
                                f"Invalid start time format: '{mapped_row.get('start_time', '')}'. Use HH:MM format (e.g., 14:30)"
                            )

                        try:
                            end_time = datetime.strptime(mapped_row.get('end_time', '').strip(), '%H:%M').time()
                        except ValueError:
                            raise ValueError(
                                f"Invalid end time format: '{mapped_row.get('end_time', '')}'. Use HH:MM format (e.g., 15:30)"
                            )

                        # Validate required text fields
                        for required_field in ['title', 'description', 'room']:
                            if not mapped_row.get(required_field) or not mapped_row[required_field].strip():
                                raise ValueError(f"Missing required field: {required_field}")

                        # Get or create convention day
                        convention_day, _ = ConventionDay.objects.get_or_create(
                            convention=convention,
                            date=parsed_date
                        )

                        # Get or create room
                        room, _ = Room.objects.get_or_create(
                            name=mapped_row['room'].strip(),
                            convention=convention
                        )

                        # Create panel
                        panel = Panel.objects.create(
                            title=mapped_row['title'].strip(),
                            description=mapped_row.get('description', '').strip(),
                            convention_day=convention_day,
                            start_time=parsed_start_time,
                            end_time=parsed_end_time,
                            room=room
                        )
                    except Exception as e:
                                return JsonResponse({'error': str(e)}, status=400)
                    except Exception as e:
                        transaction.set_rollback(True)
                        error_count += 1
                        errors.append(f"Error in row {reader.line_num}: {str(e)}")

            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} panels.')
            if error_count > 0:
                messages.error(request, f'Failed to import {error_count} panels. See details below.')
                for error in errors:
                    messages.error(request, error)

            return redirect('events:convention_detail', pk=convention.pk)
    else:
        form = CSVImportForm(initial={'convention': convention})

    # Get existing data for examples
    existing_rooms = list(Room.objects.filter(convention=convention).values_list('name', flat=True)[:5])
    existing_tags = list(Tag.objects.all().values_list('name', flat=True)[:5])  # Get all tags, not just those used in panels
    existing_hosts = list(PanelHost.objects.all().values_list('name', flat=True)[:5])  # Get all hosts, not just those used in panels

    return render(request, 'events/import_panels.html', {
        'form': form,
        'convention': convention,
        'current_convention_name': convention.name,
        'existing_rooms': existing_rooms,
        'existing_tags': existing_tags,
        'existing_hosts': existing_hosts,
        'selected_format': 'csv'
    })

@login_required
def import_panels_xlsx(request, convention_pk):
    convention = get_object_or_404(Convention, pk=convention_pk)
    
    # Get existing data for examples upfront
    existing_rooms = list(Room.objects.filter(convention=convention).values_list('name', flat=True)[:5])
    existing_tags = list(Tag.objects.all().values_list('name', flat=True)[:5])
    existing_hosts = list(PanelHost.objects.all().values_list('name', flat=True)[:5])
    
    context = {
        'convention': convention,
        'current_convention_name': convention.name,
        'existing_rooms': existing_rooms,
        'existing_tags': existing_tags,
        'existing_hosts': existing_hosts
    }
    
    if request.method == 'POST':
        form = XLSXImportForm(request.POST, request.FILES)
        if form.is_valid():
            xlsx_file = form.cleaned_data['xlsx_file']
            convention = form.cleaned_data['convention']

            # Load the workbook and get the first sheet
            try:
                wb = openpyxl.load_workbook(xlsx_file, data_only=True)
                sheet = wb.active
            except Exception as e:
                messages.error(request, f"Error reading Excel file: {str(e)}")
                context['form'] = form
                return render(request, 'events/import_panels.html', context)

            # Get headers from first row
            headers = [cell.value for cell in sheet[1] if cell.value is not None]
            if not headers:
                messages.error(request, "Excel file appears to be empty or has no headers")
                context['form'] = form
                return render(request, 'events/import_panels.html', context)

            # Map of possible column names to their canonical form
            column_mapping = {
                'title': ['title', 'Title', 'TITLE'],
                'description': ['description', 'Description', 'DESCRIPTION'],
                'date': ['date', 'Date', 'DATE'],
                'start_time': ['start time', 'Start Time', 'start_time', 'START TIME', 'START_TIME'],
                'end_time': ['end time', 'End Time', 'end_time', 'END TIME', 'END_TIME'],
                'room': ['room', 'Room', 'ROOM'],
                'tags': ['tags', 'Tags', 'TAGS'],
                'hosts': ['hosts', 'Hosts', 'HOSTS']
            }

            # Create a mapping of actual column names to canonical names
            actual_to_canonical = {}
            for canonical, possible_names in column_mapping.items():
                for name in possible_names:
                    if name in headers:
                        actual_to_canonical[name] = canonical
                        break

            # Check for missing required columns
            required_columns = ['title', 'description', 'date', 'start_time', 'end_time', 'room']
            missing_columns = [col for col in required_columns if col not in actual_to_canonical.values()]

            if missing_columns:
                messages.error(request, f"Missing required columns in Excel file: {', '.join(missing_columns)}")
                messages.info(request, "Required columns are: Title, Description, Date, Start Time, End Time, Room")
                messages.info(request, "Optional columns are: Tags, Hosts")
                context['form'] = form
                return render(request, 'events/import_panels.html', context)

            success_count = 0
            error_count = 0
            errors = []

            # Define possible date formats
            date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']

            # Process each row starting from row 2 (skip headers)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not any(cell for cell in row):
                    continue

                with transaction.atomic():
                    try:
                        # Create a dict from the row data
                        row_data = {}
                        for col_idx, cell_value in enumerate(row):
                            if col_idx < len(headers):
                                header = headers[col_idx]
                                if header in actual_to_canonical:
                                    canonical_name = actual_to_canonical[header]
                                    # Convert cell value to string, handle None values
                                    row_data[canonical_name] = str(cell_value).strip() if cell_value is not None else ''

                        # Clean and validate date value
                        date_str = row_data.get('date', '').strip()
                        if not date_str:
                            raise ValueError("Empty date value")

                        parsed_date = None
                        for date_format in date_formats:
                            try:
                                parsed_date = datetime.strptime(date_str, date_format).date()
                                break
                            except ValueError:
                                continue

                        if parsed_date is None:
                            raise ValueError(
                                f"Invalid date format: '{date_str}'. Supported formats are: YYYY-MM-DD, MM/DD/YYYY, DD/MM/YYYY, YYYY/MM/DD"
                            )

                        # Parse times
                        try:
                            start_time = datetime.strptime(row_data.get('start_time', '').strip(), '%H:%M').time()
                        except ValueError:
                            raise ValueError(
                                f"Invalid start time format: '{row_data.get('start_time', '')}'. Use HH:MM format (e.g., 14:30)"
                            )

                        try:
                            end_time = datetime.strptime(row_data.get('end_time', '').strip(), '%H:%M').time()
                        except ValueError:
                            raise ValueError(
                                f"Invalid end time format: '{row_data.get('end_time', '')}'. Use HH:MM format (e.g., 15:30)"
                            )

                        # Validate required text fields
                        for required_field in ['title', 'description', 'room']:
                            if not row_data.get(required_field) or not row_data[required_field].strip():
                                raise ValueError(f"Missing required field: {required_field}")

                        # Get or create convention day
                        convention_day, _ = ConventionDay.objects.get_or_create(
                            convention=convention,
                            date=parsed_date
                        )

                        # Get or create room
                        room, _ = Room.objects.get_or_create(
                            name=row_data['room'].strip(),
                            convention=convention
                        )

                        # Create panel
                        panel = Panel.objects.create(
                            title=row_data['title'].strip(),
                            description=row_data['description'].strip(),
                            convention_day=convention_day,
                            start_time=start_time,
                            end_time=end_time,
                            room=room
                        )

                        # Add tags
                        if row_data.get('tags'):
                            tag_names = [tag.strip() for tag in row_data['tags'].split(',')]
                            for tag_name in tag_names:
                                if tag_name:
                                    tag, _ = Tag.objects.get_or_create(name=tag_name)
                                    panel.tags.add(tag)

                        # Add hosts
                        if row_data.get('hosts'):
                            host_names = [host.strip() for host in row_data['hosts'].split(',')]
                            for index, host_name in enumerate(host_names):
                                if host_name:
                                    host, _ = PanelHost.objects.get_or_create(name=host_name)
                                    panel.host.add(host)
                                    PanelHostOrder.objects.update_or_create(
                                        panel=panel,
                                        host=host,
                                        defaults={'priority': index}
                                    )

                        success_count += 1

                    except Exception as e:
                        transaction.set_rollback(True)
                        error_count += 1
                        errors.append(f"Error in row {row_idx}: {str(e)}")

            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} panels.')
            if error_count > 0:
                messages.error(request, f'Failed to import {error_count} panels. See details below.')
                for error in errors:
                    messages.error(request, error)

            return redirect('events:convention_detail', pk=convention.pk)
    else:
        form = XLSXImportForm(initial={'convention': convention})
    
    context['form'] = form
    context['selected_format'] = 'xlsx'
    return render(request, 'events/import_panels.html', context)

def export_panels_csv(request, convention_pk):
    convention = get_object_or_404(Convention, pk=convention_pk)
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{convention.name}_schedule.csv"'
    
    # Create CSV writer with custom formatting
    writer = csv.writer(response, quoting=csv.QUOTE_ALL, lineterminator='\n')
    
    # Write header row with proper spacing
    writer.writerow([
        'Title',
        'Description',
        'Date',
        'Start Time',
        'End Time',
        'Room',
        'Tags',
        'Hosts'
    ])
    
    # Get all panels for the convention
    panels = Panel.objects.filter(
        convention_day__convention=convention
    ).select_related(
        'convention_day',
        'room'
    ).prefetch_related(
        'tags',
        'host'
    ).order_by('convention_day__date', 'start_time')
    
    # Write panel data
    for panel in panels:
        # Get tags and hosts as comma-separated strings with proper spacing
        tags = ', '.join(tag.name.strip() for tag in panel.tags.all())
        hosts = ', '.join(host.name.strip() for host in panel.host.all().order_by('panelhostorder__priority'))
        
        # Clean and format the description
        description = panel.description.strip().replace('\n', ' ').replace('\r', '')
        
        writer.writerow([
            panel.title.strip(),
            description,
            panel.convention_day.date.strftime('%Y-%m-%d'),
            panel.start_time.strftime('%H:%M'),
            panel.end_time.strftime('%H:%M'),
            panel.room.name.strip() if panel.room else '',
            tags,
            hosts
        ])
    
    return response


def convention_ical_feed(request, pk):
    convention = get_object_or_404(Convention, pk=pk)
    days = convention.days.all().order_by('date')
    cal = icalendar.Calendar()
    cal.add('prodid', '-//FurConnect//Convention Schedule//EN')
    cal.add('version', '2.0')
    cal.add('X-WR-CALNAME', convention.name)
    # Try to get the convention's timezone from its location, fallback to UTC
    import pytz
    from timezonefinder import TimezoneFinder
    import geopy.geocoders
    tz_name = 'UTC'
    try:
        geolocator = geopy.geocoders.Nominatim(user_agent="furconnect-ical")
        location = geolocator.geocode(convention.location)
        if location:
            tf = TimezoneFinder()
            tz_name = tf.timezone_at(lng=location.longitude, lat=location.latitude) or 'UTC'
    except Exception:
        tz_name = 'UTC'
    cal.add('X-WR-TIMEZONE', tz_name)
    tz = pytz.timezone(tz_name)
    for day in days:
        panels = day.panels.filter(cancelled=False).order_by('start_time')
        for panel in panels:
            event = icalendar.Event()
            event.add('summary', panel.title or "Untitled Event")
            event.add('description', panel.description or "")
            room_name = panel.room.name if panel.room else ""
            event.add('location', f"{convention.name} - {room_name}" if room_name else convention.name)
            start_datetime = datetime.combine(day.date, panel.start_time)
            end_datetime = datetime.combine(day.date, panel.end_time)
            # If naive, localize; if aware, convert to convention tz
            if start_datetime.tzinfo is None:
                start_datetime = tz.localize(start_datetime)
            else:
                start_datetime = start_datetime.astimezone(tz)
            if end_datetime.tzinfo is None:
                end_datetime = tz.localize(end_datetime)
            else:
                end_datetime = end_datetime.astimezone(tz)
            if end_datetime < start_datetime:
                end_datetime += timedelta(days=1)
            event.add('dtstart', start_datetime)
            event.add('dtend', end_datetime)
            event.add('dtstamp', timezone.now().astimezone(tz))
            # Add a unique identifier
            event.add('uid', f"panel-{panel.pk}@furconnect")
            cal.add_component(event)
    response = HttpResponse(cal.to_ical(), content_type='text/calendar')
    response['Content-Disposition'] = f'inline; filename="{convention.name}_schedule.ics"'
    return response

# Batch host details AJAX endpoint
    @require_GET
    def batch_host_details(request):
            try:
                host_ids = request.GET.getlist('host_ids[]')
                if not host_ids:
                    return JsonResponse({'error': 'No host IDs provided'}, status=400)

                hosts = PanelHost.objects.filter(pk__in=host_ids).prefetch_related('panels')
                hosts_data = []
                for host in hosts:
                    panels_data = []
                    for panel in host.panels.all():
                        panels_data.append({
                            'id': panel.pk,
                            'title': panel.title,
                            'date': panel.convention_day.date.strftime('%Y-%m-%d'),
                            'start_time': panel.start_time.strftime('%H:%M'),
                            'end_time': panel.end_time.strftime('%H:%M'),
                            'room': panel.room.name if panel.room else 'TBD'
                        })
                    hosts_data.append({
                        'id': host.pk,
                        'name': host.name,
                        'image': host.image,
                        'panels': panels_data,
                        'panels_count': len(panels_data),
                    })
                return JsonResponse({'hosts': hosts_data})
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=400)

@login_required
def download_csv_template(request, convention_pk):
    convention = get_object_or_404(Convention, pk=convention_pk)
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{convention.name}_panels_template.csv"'
    
    # Create CSV writer
    writer = csv.writer(response, quoting=csv.QUOTE_ALL, lineterminator='\n')
    
    # Write header row
    writer.writerow([
        'title',
        'description', 
        'date',
        'start_time',
        'end_time',
        'room',
        'tags',
        'hosts'
    ])
    
    # Get example data
    existing_rooms = list(Room.objects.filter(convention=convention).values_list('name', flat=True)[:1]) or ['Main Hall']
    existing_tags = list(Tag.objects.all().values_list('name', flat=True)[:2]) or ['Art', 'Workshop']
    existing_hosts = list(PanelHost.objects.all().values_list('name', flat=True)[:2]) or ['John Doe', 'Jane Smith']
    
    # Write example row
    writer.writerow([
        'Furry Art Workshop',
        'Learn to draw furry art',
        '2024-03-15',
        '10:00',
        '12:00',
        existing_rooms[0],
        ', '.join(existing_tags),
        ', '.join(existing_hosts)
    ])
    
    return response

@login_required
def download_xlsx_template(request, convention_pk):
    convention = get_object_or_404(Convention, pk=convention_pk)
    
    # Create workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Panels'
    
    # Write headers
    headers = ['title', 'description', 'date', 'start_time', 'end_time', 'room', 'tags', 'hosts']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    
    # Get example data
    existing_rooms = list(Room.objects.filter(convention=convention).values_list('name', flat=True)[:1]) or ['Main Hall']
    existing_tags = list(Tag.objects.all().values_list('name', flat=True)[:2]) or ['Art', 'Workshop']
    existing_hosts = list(PanelHost.objects.all().values_list('name', flat=True)[:2]) or ['John Doe', 'Jane Smith']
    
    # Write example row
    example_data = [
        'Furry Art Workshop',
        'Learn to draw furry art',
        '2024-03-15',
        '10:00',
        '12:00',
        existing_rooms[0],
        ', '.join(existing_tags),
        ', '.join(existing_hosts)
    ]
    for col_num, value in enumerate(example_data, 1):
        sheet.cell(row=2, column=col_num, value=value)
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{convention.name}_panels_template.xlsx"'
    
    wb.save(response)
    return response
@login_required
def download_csv_template(request, convention_pk):
    convention = get_object_or_404(Convention, pk=convention_pk)
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{convention.name}_panels_template.csv"'
    
    # Create CSV writer
    writer = csv.writer(response, quoting=csv.QUOTE_ALL, lineterminator='\n')
    
    # Write header row
    writer.writerow([
        'title',
        'description', 
        'date',
        'start_time',
        'end_time',
        'room',
        'tags',
        'hosts'
    ])
    
    # Get example data
    existing_rooms = list(Room.objects.filter(convention=convention).values_list('name', flat=True)[:1]) or ['Main Hall']
    existing_tags = list(Tag.objects.all().values_list('name', flat=True)[:2]) or ['Art', 'Workshop']
    existing_hosts = list(PanelHost.objects.all().values_list('name', flat=True)[:2]) or ['John Doe', 'Jane Smith']
    
    # Write example row
    writer.writerow([
        'Furry Art Workshop',
        'Learn to draw furry art',
        '2024-03-15',
        '10:00',
        '12:00',
        existing_rooms[0],
        ', '.join(existing_tags),
        ', '.join(existing_hosts)
    ])
    
    return response

@login_required
def download_xlsx_template(request, convention_pk):
    convention = get_object_or_404(Convention, pk=convention_pk)
    
    # Create workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Panels'
    
    # Write headers
    headers = ['title', 'description', 'date', 'start_time', 'end_time', 'room', 'tags', 'hosts']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    
    # Get example data
    existing_rooms = list(Room.objects.filter(convention=convention).values_list('name', flat=True)[:1]) or ['Main Hall']
    existing_tags = list(Tag.objects.all().values_list('name', flat=True)[:2]) or ['Art', 'Workshop']
    existing_hosts = list(PanelHost.objects.all().values_list('name', flat=True)[:2]) or ['John Doe', 'Jane Smith']
    
    # Write example row
    example_data = [
        'Furry Art Workshop',
        'Learn to draw furry art',
        '2024-03-15',
        '10:00',
        '12:00',
        existing_rooms[0],
        ', '.join(existing_tags),
        ', '.join(existing_hosts)
    ]
    for col_num, value in enumerate(example_data, 1):
        sheet.cell(row=2, column=col_num, value=value)
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{convention.name}_panels_template.xlsx"'
    
    wb.save(response)
    return response
